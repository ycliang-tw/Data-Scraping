## import necessery module
# import urllib.request as req
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = "標題"
ws['B1'] = "作者"
ws['C1'] = "內文"
ws['D1'] = "連結"

## set up requests
url = "https://www.ptt.cc/bbs/VAPE/index.html"
url2 = "https://www.ptt.cc/ask/over18?from=%2Fbbs%2FVAPE%2Findex.html"

headers = {	"User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36"}
payload = {	"from": "/bbs/VAPE/index.html", "yes": "yes"}

r = requests.Session()
r1 = r.post(url2, payload)

for i in range(100):
	r2 = r.get(url, headers=headers)
	soup = BeautifulSoup(r2.text, "html.parser")
	r_ent = soup.find_all(class_="r-ent")
	pre = soup.select("div.btn-group.btn-group-paging a")
	url = "https://www.ptt.cc" + pre[1]["href"]
	for entry in r_ent:
		title = entry.find(class_="title").text.strip()
		if "買賣" in title:
			url2 = "https://www.ptt.cc" + entry.find('a')['href']
			author = entry.find(class_="author").text.strip()
			r3 = r.get(url2)
			soup2 = BeautifulSoup(r3.text, "html.parser")
			content = soup2.find(id="main-content").text
			target = '※ 發信站: 批踢踢實業坊(ptt.cc),'
			content = content.split(target)
			try:
				time = soup2.select(".article-metaline")[2].text
				content = content[0].split(time)
			except Exception as e:
				print(e)
			print(url2)
			print(author)
			print(title)
			print(content[1])
			ws.append([title, author, content[1], url2])

wb.save("PTT_VAPE.xlsx")
